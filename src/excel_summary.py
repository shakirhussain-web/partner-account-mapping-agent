import os
from collections import defaultdict
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from format import to_usd, _quarter_bounds, _bucket_deal, _collapse_opps

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "agent-output")

REGIONS = ["AMER", "EMEA", "APAC", "LATAM"]

REGION_MAP = {
    "americas": "AMER", "amer": "AMER", "na": "AMER", "north america": "AMER",
    "emea": "EMEA", "europe": "EMEA",
    "apac": "APAC", "asia pacific": "APAC",
    "latam": "LATAM", "latin america": "LATAM",
}


def _normalize_region(region):
    if not region:
        return "Unknown"
    return REGION_MAP.get(region.lower().strip(), region)


def _aggregate_partner(details, subscriptions, bookings, open_pipeline):
    row = {}

    # Partner details
    if details:
        tiers = [d.get("CHANNEL_CATEGORY") for d in details if d.get("CHANNEL_CATEGORY")]
        tier_order = {"Premier": 0, "Advanced": 1, "Qualified": 2}
        tiers.sort(key=lambda t: tier_order.get(t, 99))
        row["tier"] = tiers[0] if tiers else "N/A"

        agreements = sorted(set(
            str(d.get("SIGNED_AGREEMENT") or "") for d in details if d.get("SIGNED_AGREEMENT")
        ))
        row["agreement"] = ", ".join(agreements) if agreements else "N/A"

        owners = sorted(set(
            str(d.get("ACCOUNT_OWNER") or "") for d in details if d.get("ACCOUNT_OWNER")
        ))
        row["owners"] = ", ".join(owners) if owners else "N/A"
    else:
        row["tier"] = "N/A"
        row["agreement"] = "N/A"
        row["owners"] = "N/A"

    # Book of Business by region
    bob_by_region = defaultdict(lambda: {"customers": set(), "arr": 0})
    for sub in (subscriptions or []):
        cust = sub.get("RESELLERCUSTOMER_ACCOUNTNAME")
        if not cust:
            continue
        region = _normalize_region(sub.get("RESELLERCUSTOMER_REGION"))
        arr = to_usd(sub.get("RESELLERCUSTOMER_ARR", 0), sub.get("RESELLERCUSTOMER_CURRENCY", "USD"))
        bob_by_region[region]["customers"].add(cust)
        bob_by_region[region]["arr"] = max(bob_by_region[region]["arr"],
                                           bob_by_region[region]["arr"])
    # Recalculate properly: unique customers per region, sum ARR per customer
    bob_customers = defaultdict(lambda: {"regions": set(), "arr": 0, "industry": None})
    for sub in (subscriptions or []):
        cust = sub.get("RESELLERCUSTOMER_ACCOUNTNAME")
        if not cust:
            continue
        region = _normalize_region(sub.get("RESELLERCUSTOMER_REGION"))
        arr = to_usd(sub.get("RESELLERCUSTOMER_ARR", 0), sub.get("RESELLERCUSTOMER_CURRENCY", "USD"))
        bob_customers[cust]["regions"].add(region)
        bob_customers[cust]["arr"] = arr
        if sub.get("RESELLERCUSTOMER_INDUSTRY"):
            bob_customers[cust]["industry"] = sub["RESELLERCUSTOMER_INDUSTRY"]

    bob_region_agg = defaultdict(lambda: {"customers": 0, "arr": 0})
    for cust, data in bob_customers.items():
        for region in data["regions"]:
            bob_region_agg[region]["customers"] += 1
            bob_region_agg[region]["arr"] += data["arr"]

    for r in REGIONS:
        row[f"bob_{r.lower()}_customers"] = bob_region_agg.get(r, {}).get("customers", 0)
        row[f"bob_{r.lower()}_arr"] = bob_region_agg.get(r, {}).get("arr", 0)

    row["bob_total_customers"] = len(bob_customers)
    row["bob_total_arr"] = sum(c["arr"] for c in bob_customers.values())

    # Book of Business by industry
    bob_by_industry = defaultdict(lambda: 0)
    for cust, data in bob_customers.items():
        industry = data.get("industry") or "Unknown"
        bob_by_industry[industry] += data["arr"]
    top_bob_industries = sorted(bob_by_industry.items(), key=lambda x: x[1], reverse=True)[:5]
    for i in range(5):
        if i < len(top_bob_industries):
            ind, arr = top_bob_industries[i]
            row[f"bob_industry_{i+1}"] = f"{ind} - ${arr:,.0f}"
        else:
            row[f"bob_industry_{i+1}"] = ""

    # Bookings by region
    bookings_by_region = defaultdict(lambda: 0)
    total_bookings = 0
    sourced = 0
    influenced = 0
    deal_count = 0
    for b in (bookings or []):
        region = _normalize_region(b.get("REGION"))
        arr = b.get("BOOKINGS", 0) or 0
        bookings_by_region[region] += arr
        total_bookings += arr
        deal_count += 1
        if b.get("SOURCED_INFLUENCED") == "Partner Sourced":
            sourced += arr
        elif b.get("SOURCED_INFLUENCED") == "Partner Influenced":
            influenced += arr

    for r in REGIONS:
        row[f"bookings_{r.lower()}"] = bookings_by_region.get(r, 0)

    row["bookings_total"] = total_bookings
    row["bookings_sourced"] = sourced
    row["bookings_influenced"] = influenced
    row["bookings_deals"] = deal_count

    # Bookings by industry
    bookings_by_industry = defaultdict(lambda: 0)
    for b in (bookings or []):
        industry = b.get("INDUSTRY") or "Unknown"
        bookings_by_industry[industry] += b.get("BOOKINGS", 0) or 0
    top_booking_industries = sorted(bookings_by_industry.items(), key=lambda x: x[1], reverse=True)[:5]
    for i in range(5):
        if i < len(top_booking_industries):
            ind, arr = top_booking_industries[i]
            row[f"bookings_industry_{i+1}"] = f"{ind} - ${arr:,.0f}"
        else:
            row[f"bookings_industry_{i+1}"] = ""

    # Open Pipeline
    today = date.today()
    cq_start, cq1_start, cq1_end, fy, fq = _quarter_bounds(today)
    cq_arr = 0
    cq1_arr = 0
    total_pipe = 0
    pipe_deals = 0

    seen_opps = set()
    for p in (open_pipeline or []):
        opp_id = p.get("CRM_OPPORTUNITY_ID")
        product = str(p.get("PRODUCT", "") or "")
        if product.lower() != "total booking":
            continue
        if opp_id in seen_opps:
            continue
        seen_opps.add(opp_id)

        arr = p.get("PRODUCT_ARR_USD", 0) or 0
        bucket = _bucket_deal(p.get("CLOSEDATE"), cq_start, cq1_start, cq1_end)
        total_pipe += arr
        pipe_deals += 1
        if bucket == "cq":
            cq_arr += arr
        elif bucket == "cq1":
            cq1_arr += arr

    row["pipe_cq"] = cq_arr
    row["pipe_cq1"] = cq1_arr
    row["pipe_total"] = total_pipe
    row["pipe_deals"] = pipe_deals

    # Pipeline by industry (using collapsed opps to avoid double-counting)
    pipe_by_industry = defaultdict(lambda: 0)
    for p in (open_pipeline or []):
        product = str(p.get("PRODUCT", "") or "")
        if product.lower() != "total booking":
            continue
        opp_id = p.get("CRM_OPPORTUNITY_ID")
        industry = p.get("INDUSTRY") or "Unknown"
        pipe_by_industry[(opp_id, industry)] = p.get("PRODUCT_ARR_USD", 0) or 0
    industry_totals = defaultdict(lambda: 0)
    for (opp_id, industry), arr in pipe_by_industry.items():
        industry_totals[industry] += arr
    top_pipe_industries = sorted(industry_totals.items(), key=lambda x: x[1], reverse=True)[:5]
    for i in range(5):
        if i < len(top_pipe_industries):
            ind, arr = top_pipe_industries[i]
            row[f"pipe_industry_{i+1}"] = f"{ind} - ${arr:,.0f}"
        else:
            row[f"pipe_industry_{i+1}"] = ""

    return row


def generate_excel(partners_data):
    today = date.today()
    _, _, _, fy, fq = _quarter_bounds(today)
    nq = fq + 1 if fq < 4 else 1
    nfy = fy if fq < 4 else fy + 1
    cq_label = f"FY{fy}Q{fq}"
    cq1_label = f"FY{nfy}Q{nq}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Partner Summary"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="B0B0B0"),
    )
    money_fmt = '#,##0'

    headers = [
        "Partner Name",
        "Partner Tier", "Agreement", "Account Owner(s)",
        "BoB AMER #", "BoB AMER ARR",
        "BoB EMEA #", "BoB EMEA ARR",
        "BoB APAC #", "BoB APAC ARR",
        "BoB LATAM #", "BoB LATAM ARR",
        "BoB Total #", "BoB Total ARR",
        "BoB Industry 1", "BoB Industry 2", "BoB Industry 3", "BoB Industry 4", "BoB Industry 5",
        "Bookings AMER", "Bookings EMEA", "Bookings APAC", "Bookings LATAM",
        "Bookings Total", "Bookings Sourced", "Bookings Influenced", "Bookings # Deals",
        "Bookings Industry 1", "Bookings Industry 2", "Bookings Industry 3", "Bookings Industry 4", "Bookings Industry 5",
        f"Pipeline {cq_label}", f"Pipeline {cq1_label}", "Pipeline Total", "Pipeline # Deals",
        "Pipeline Industry 1", "Pipeline Industry 2", "Pipeline Industry 3", "Pipeline Industry 4", "Pipeline Industry 5",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (partner_name, data) in enumerate(partners_data, 2):
        agg = _aggregate_partner(
            data.get("details"),
            data.get("subscriptions"),
            data.get("bookings"),
            data.get("open_pipeline"),
        )

        values = [
            partner_name,
            agg["tier"], agg["agreement"], agg["owners"],
            agg["bob_amer_customers"], agg["bob_amer_arr"],
            agg["bob_emea_customers"], agg["bob_emea_arr"],
            agg["bob_apac_customers"], agg["bob_apac_arr"],
            agg["bob_latam_customers"], agg["bob_latam_arr"],
            agg["bob_total_customers"], agg["bob_total_arr"],
            agg["bob_industry_1"], agg["bob_industry_2"], agg["bob_industry_3"], agg["bob_industry_4"], agg["bob_industry_5"],
            agg["bookings_amer"], agg["bookings_emea"],
            agg["bookings_apac"], agg["bookings_latam"],
            agg["bookings_total"], agg["bookings_sourced"],
            agg["bookings_influenced"], agg["bookings_deals"],
            agg["bookings_industry_1"], agg["bookings_industry_2"], agg["bookings_industry_3"], agg["bookings_industry_4"], agg["bookings_industry_5"],
            agg["pipe_cq"], agg["pipe_cq1"],
            agg["pipe_total"], agg["pipe_deals"],
            agg["pipe_industry_1"], agg["pipe_industry_2"], agg["pipe_industry_3"], agg["pipe_industry_4"], agg["pipe_industry_5"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if isinstance(val, (int, float)) and col >= 5:
                cell.number_format = money_fmt

    industry_cols = set()
    for col in range(1, len(headers) + 1):
        h = headers[col - 1]
        letter = ws.cell(row=1, column=col).column_letter
        if "Industry" in h:
            ws.column_dimensions[letter].width = 35
            industry_cols.add(col)
        else:
            ws.column_dimensions[letter].width = 16
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 25

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "B2"

    filename = f"Partner_Summary_{today.isoformat()}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(filepath)
    return filepath
