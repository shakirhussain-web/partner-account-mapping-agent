#!/usr/bin/env python3
import json
import os
import sys
import openpyxl
from snowflake_conn import get_connection, execute_query, close
from queries import (reseller_subscriptions_query, partner_bookings_query,
                     partner_details_query, partner_open_pipeline_query,
                     sourced_pipeline_query)
from format import format_partner_report, _fiscal_quarter
from pdf_report import generate_pdf
from excel_summary import generate_excel

ALIASES_PATH = os.path.join(os.path.dirname(__file__), "aliases.json")
INPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "agent-input")


def load_aliases():
    try:
        with open(ALIASES_PATH) as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def resolve_names(partner_name, aliases):
    for key, names in aliases.items():
        if partner_name.lower() == key.lower():
            return key, names
        if partner_name.lower() in [n.lower() for n in names]:
            return key, names
    return partner_name, [partner_name]


def extract_partners_from_excel():
    raw_names = set()
    for fname in os.listdir(INPUT_DIR):
        if not fname.endswith(".xlsx"):
            continue
        wb = openpyxl.load_workbook(os.path.join(INPUT_DIR, fname), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip().lower() if c else "" for c in rows[0]]
            name_col = None
            for i, h in enumerate(header):
                if "partner" in h and "name" in h:
                    name_col = i
                    break
            if name_col is None:
                continue
            for row in rows[1:]:
                val = row[name_col] if name_col < len(row) else None
                if val and str(val).strip():
                    raw_names.add(str(val).strip())
        wb.close()
    return raw_names


def deduplicate(raw_names, aliases):
    seen = set()
    partners = []
    for name in sorted(raw_names):
        display_name, search_names = resolve_names(name, aliases)
        key = display_name.lower()
        if key not in seen:
            seen.add(key)
            partners.append((display_name, search_names))
    return partners


def main():
    fmt = "pdf"
    if "--format" in sys.argv:
        idx = sys.argv.index("--format")
        if idx + 1 < len(sys.argv):
            fmt = sys.argv[idx + 1].lower()

    if fmt not in ("pdf", "xlsx", "both"):
        print("Usage: python3 batch_partner_lookup.py [--format pdf|xlsx|both]")
        sys.exit(1)

    extra_names = []
    if "--add" in sys.argv:
        idx = sys.argv.index("--add")
        extra_names = [n.strip() for n in sys.argv[idx + 1:] if not n.startswith("--")]

    print("Reading Excel files from agent-input/...")
    raw_names = extract_partners_from_excel()
    if extra_names:
        raw_names.update(extra_names)
        print(f"  Added extra partners: {', '.join(extra_names)}")
    print(f"  Found {len(raw_names)} raw partner names")

    aliases = load_aliases()
    partners = deduplicate(raw_names, aliases)
    print(f"  Deduplicated to {len(partners)} unique partners")
    print(f"  Output format: {fmt}\n")

    from datetime import date as dt_date, timedelta
    today = dt_date.today()
    fy, fq, cq_start, _ = _fiscal_quarter(today)
    fy1, fq1, cqm1_start, _ = _fiscal_quarter(cq_start - timedelta(days=1))
    fy2, fq2, _, _ = _fiscal_quarter(cqm1_start - timedelta(days=1))
    fiscal_quarters = [f"FY{fy2}Q{fq2}", f"FY{fy1}Q{fq1}", f"FY{fy}Q{fq}"]
    print(f"  Sourced pipeline quarters: {', '.join(fiscal_quarters)}\n")

    all_partner_data = []

    for i, (display_name, search_names) in enumerate(partners, 1):
        label = display_name
        if len(search_names) > 1:
            label += f" (incl. {', '.join(n for n in search_names if n.lower() != display_name.lower())})"

        print(f"[{i}/{len(partners)}] {label}")

        try:
            details = execute_query(partner_details_query(search_names))
            subs = execute_query(reseller_subscriptions_query(search_names))
            bookings = execute_query(partner_bookings_query(search_names))
            open_pipe = execute_query(partner_open_pipeline_query(search_names))
            sourced = execute_query(sourced_pipeline_query(search_names, fiscal_quarters))

            print(f"  Details: {len(details)}, Subs: {len(subs)}, Bookings: {len(bookings)}, Pipeline: {len(open_pipe)}, Sourced: {len(sourced)}")

            if fmt in ("pdf", "both"):
                report = format_partner_report(display_name, subs, bookings,
                                               details=details, open_pipeline=open_pipe,
                                               sourced_pipeline=sourced)
                print(report)
                pdf_path = generate_pdf(display_name, subs, bookings,
                                        details=details, open_pipeline=open_pipe,
                                        sourced_pipeline=sourced)
                print(f"  PDF saved: {pdf_path}")

            all_partner_data.append((display_name, {
                "details": details,
                "subscriptions": subs,
                "bookings": bookings,
                "open_pipeline": open_pipe,
                "sourced_pipeline": sourced,
            }))
            print()
        except Exception as e:
            print(f"  ERROR: {e}\n")

    if fmt in ("xlsx", "both"):
        print("Generating Excel summary...")
        xlsx_path = generate_excel(all_partner_data)
        print(f"  Excel saved: {xlsx_path}")

    close()
    print("\nDone!")


if __name__ == "__main__":
    main()
